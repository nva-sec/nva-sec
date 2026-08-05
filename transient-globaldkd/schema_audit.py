#!/usr/bin/env python3
"""Schema-only audit for Zenodo GlobalDKD.

This program intentionally does not read reference-value cells, construct
albumin labels, select wavelengths, fit models, or calculate performance.
"""
from __future__ import annotations

import hashlib
import io
import json
import re
import sys
import tempfile
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from xml.etree import ElementTree

import h5py
from openpyxl import load_workbook
from scipy.io import whosmat

EXPECTED_MD5 = "8d5d4ba0e0578d0e39e76e1f1514a841"
EXPECTED = {
    "australia_donors": 177,
    "spain_donors": 61,
    "valencia_donors": 46,
    "madrid_donors": 15,
}


def md5sum(path: Path) -> str:
    digest = hashlib.md5()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def canonical_australia(filename: str) -> str:
    donor = re.sub(r"\.\d+$", "", filename)
    donor = re.sub(r"_V\d+$", "", donor, flags=re.IGNORECASE)
    # The archive inventory exposes HEA16_2 as a rerun of HEA16.
    if donor.upper() == "HEA16_2":
        donor = "HEA16"
    return donor.upper()


def canonical_spain(filename: str) -> str:
    donor = re.sub(r"\.sp$", "", filename, flags=re.IGNORECASE)
    donor = re.sub(r"_(?:R\d+|CONC)$", "", donor, flags=re.IGNORECASE)
    return donor.upper()


def mat_schema(payload: bytes, name: str) -> dict:
    header = payload[:128].decode("latin-1", errors="replace").strip("\x00 ")
    with tempfile.NamedTemporaryFile(suffix=".mat") as tmp:
        tmp.write(payload)
        tmp.flush()
        try:
            variables = [
                {"name": var, "shape": list(shape), "class": cls}
                for var, shape, cls in whosmat(tmp.name)
            ]
            return {
                "path": name,
                "storage": "MAT v4/v5/v7 (scipy whosmat)",
                "header": header,
                "variables": variables,
            }
        except (NotImplementedError, ValueError, OSError) as scipy_error:
            datasets = []

            def visitor(dataset_name, obj):
                if isinstance(obj, h5py.Dataset):
                    datasets.append(
                        {
                            "name": dataset_name,
                            "shape": list(obj.shape),
                            "dtype": str(obj.dtype),
                        }
                    )

            try:
                with h5py.File(tmp.name, "r") as handle:
                    handle.visititems(visitor)
                return {
                    "path": name,
                    "storage": "MAT v7.3/HDF5",
                    "header": header,
                    "datasets": datasets,
                }
            except OSError as hdf_error:
                return {
                    "path": name,
                    "storage": "unreadable",
                    "header": header,
                    "scipy_error": repr(scipy_error),
                    "hdf5_error": repr(hdf_error),
                }


def mlx_schema(payload: bytes, name: str, output_dir: Path) -> dict:
    result = {"path": name, "container": None, "members": [], "code_nodes": 0}
    try:
        with zipfile.ZipFile(io.BytesIO(payload)) as live:
            result["container"] = "zip"
            result["members"] = sorted(live.namelist())
            code_blocks = []
            for member in live.namelist():
                if not member.lower().endswith(".xml"):
                    continue
                try:
                    root = ElementTree.fromstring(live.read(member))
                except ElementTree.ParseError:
                    continue
                for node in root.iter():
                    tag = node.tag.rsplit("}", 1)[-1].lower()
                    if tag in {"mcode", "code", "input"}:
                        text = "".join(node.itertext()).strip()
                        if text:
                            code_blocks.append(text)
            result["code_nodes"] = len(code_blocks)
            if code_blocks:
                safe = re.sub(r"[^A-Za-z0-9_.-]+", "_", name)
                (output_dir / f"{safe}.code.txt").write_text(
                    "\n\n% ---- code node ----\n\n".join(code_blocks),
                    encoding="utf-8",
                )
    except zipfile.BadZipFile:
        result["container"] = "not-zip"
    return result


def workbook_schema(payload: bytes, name: str) -> dict:
    # Header/schema only: no reference-value rows are read.
    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=False)
    sheets = []
    for sheet in workbook.worksheets:
        header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        sheets.append(
            {
                "title": sheet.title,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "header": [str(value) if value is not None else None for value in header],
            }
        )
    workbook.close()
    return {"path": name, "sheets": sheets}


def main() -> None:
    archive = Path(sys.argv[1] if len(sys.argv) > 1 else "data/GlobalDKD.zip")
    output_dir = Path(sys.argv[2] if len(sys.argv) > 2 else "outputs/schema-audit")
    output_dir.mkdir(parents=True, exist_ok=True)

    observed_md5 = md5sum(archive)
    if observed_md5 != EXPECTED_MD5:
        raise SystemExit(f"archive MD5 mismatch: {observed_md5}")

    with zipfile.ZipFile(archive) as bundle:
        files = sorted(info.filename for info in bundle.infolist() if not info.is_dir())
        suffix_counts = Counter(Path(name).suffix.lower() or "<none>" for name in files)

        australian_spectra = [
            name
            for name in files
            if "/australia/" in f"/{name.lower()}"
            and re.search(r"\.\d+$", Path(name).name)
        ]
        australia_groups = defaultdict(list)
        for name in australian_spectra:
            australia_groups[canonical_australia(Path(name).name)].append(name)

        spanish_spectra = [
            name
            for name in files
            if "/spain/" in f"/{name.lower()}" and name.lower().endswith(".sp")
            and "blank" not in Path(name).name.lower()
            and "blanco" not in Path(name).name.lower()
        ]
        spain_groups = defaultdict(list)
        site_counts = Counter()
        for name in spanish_spectra:
            donor = canonical_spain(Path(name).name)
            spain_groups[donor].append(name)
            lowered = name.lower()
            if "/valencia/" in f"/{lowered}":
                site_counts["Valencia"] += 1
            elif "/madrid/" in f"/{lowered}":
                site_counts["Madrid"] += 1
            else:
                site_counts["Unknown"] += 1

        mat_files = [name for name in files if name.lower().endswith(".mat")]
        mlx_files = [name for name in files if name.lower().endswith(".mlx")]
        xlsx_files = [name for name in files if name.lower().endswith(".xlsx")]

        report = {
            "phase": "schema-only; no outcome values or models",
            "archive": {
                "path": str(archive),
                "bytes": archive.stat().st_size,
                "md5": observed_md5,
                "member_count": len(files),
                "suffix_counts": dict(sorted(suffix_counts.items())),
            },
            "cohort_inventory": {
                "australia_raw_spectra": len(australian_spectra),
                "australia_unique_donor_ids": len(australia_groups),
                "australia_replicate_histogram": dict(
                    sorted(Counter(map(len, australia_groups.values())).items())
                ),
                "spain_raw_spectra_excluding_blanks": len(spanish_spectra),
                "spain_unique_donor_ids": len(spain_groups),
                "spain_site_counts": dict(site_counts),
                "excluded_spanish_blanks": [
                    name for name in files
                    if "/spain/" in f"/{name.lower()}"
                    and name.lower().endswith(".sp")
                    and ("blank" in Path(name).name.lower() or "blanco" in Path(name).name.lower())
                ],
            },
            "matlab": [mat_schema(bundle.read(name), name) for name in mat_files],
            "live_scripts": [
                mlx_schema(bundle.read(name), name, output_dir) for name in mlx_files
            ],
            "workbooks": [
                workbook_schema(bundle.read(name), name) for name in xlsx_files
            ],
        }

    observed = {
        "australia_donors": report["cohort_inventory"]["australia_unique_donor_ids"],
        "spain_donors": report["cohort_inventory"]["spain_unique_donor_ids"],
        "valencia_donors": report["cohort_inventory"]["spain_site_counts"].get("Valencia", 0),
        "madrid_donors": report["cohort_inventory"]["spain_site_counts"].get("Madrid", 0),
    }
    report["count_assertions"] = {
        key: {"expected": EXPECTED[key], "observed": observed[key], "pass": observed[key] == EXPECTED[key]}
        for key in EXPECTED
    }

    (output_dir / "schema-audit.json").write_text(
        json.dumps(report, indent=2, sort_keys=True), encoding="utf-8"
    )
    (output_dir / "archive-members.txt").write_text(
        "\n".join(files) + "\n", encoding="utf-8"
    )

    lines = [
        "# GlobalDKD schema audit",
        "",
        "**Phase:** schema only; no outcome values, labels, filters, models, or performance.",
        "",
        f"- Archive MD5: `{observed_md5}`",
        f"- Archive members: {len(files)}",
        f"- Australian raw spectra: {len(australian_spectra)}",
        f"- Australian inferred donors: {observed['australia_donors']} / {EXPECTED['australia_donors']}",
        f"- Spanish inferred donors: {observed['spain_donors']} / {EXPECTED['spain_donors']}",
        f"- Valencia: {observed['valencia_donors']} / {EXPECTED['valencia_donors']}",
        f"- Madrid: {observed['madrid_donors']} / {EXPECTED['madrid_donors']}",
        f"- MAT files: {len(mat_files)}; MLX files: {len(mlx_files)}; workbooks: {len(xlsx_files)}",
        "",
        "Detailed variable shapes and container inventories are in `schema-audit.json`.",
    ]
    (output_dir / "README.md").write_text("\n".join(lines) + "\n", encoding="utf-8")

    failed = [key for key, check in report["count_assertions"].items() if not check["pass"]]
    if failed:
        raise SystemExit(f"donor-count assertion failed: {', '.join(failed)}")
    if any(item["storage"] == "unreadable" for item in report["matlab"]):
        raise SystemExit("one or more MAT files were unreadable")
    print(json.dumps({"status": "PASS", "counts": observed, "md5": observed_md5}, indent=2))


if __name__ == "__main__":
    main()
